#!/usr/bin/env python3
"""
מעקב אימוני SUP — גרמין קונקט → אקסל (אוטומטי מלא)
הרצה: python garmin_sup_tracker.py
הגדרת מיקומים (פעם ראשונה): python garmin_sup_tracker.py --setup
"""

import json
import math
import os
import sys
import getpass
import argparse
import webbrowser
import tempfile
import smtplib
import time
import urllib.request
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime, timedelta
from copy import copy

import numpy as np
import openpyxl
import garminconnect
try:
    from winotify import Notification, audio
    TOAST_AVAILABLE = True
except ImportError:
    TOAST_AVAILABLE = False

DAYS_HEB = {0:"שני", 1:"שלישי", 2:"רביעי", 3:"חמישי", 4:"שישי", 5:"שבת", 6:"ראשון"}

# ── בדיקת חיבור לאינטרנט ─────────────────────────────────────────────────────
def wait_for_internet(max_wait_min=30, check_interval_sec=60):
    """ממתין לחיבור אינטרנט — עד max_wait_min דקות"""
    def is_connected():
        try:
            urllib.request.urlopen("https://www.google.com", timeout=5)
            return True
        except Exception:
            return False

    if is_connected():
        return True

    print(f"אין חיבור לאינטרנט — ממתין (עד {max_wait_min} דקות)...")
    waited = 0
    while waited < max_wait_min * 60:
        time.sleep(check_interval_sec)
        waited += check_interval_sec
        if is_connected():
            print(f"  חיבור נמצא לאחר {waited//60} דקות")
            return True
        print(f"  עדיין ממתין... ({waited//60}/{max_wait_min} דקות)")

    print("לא נמצא חיבור אינטרנט — הסקריפט מסתיים.")
    return False

# ── קבצי הגדרות ──────────────────────────────────────────────────────────────
_BASE = os.path.dirname(os.path.abspath(__file__))          # tracker/
_PROJ = os.path.dirname(_BASE)                              # sup-challenge/

EXCEL_FILE     = os.path.join(_PROJ, "ניתוח_אימוני_SUP_v5.xlsx")
CONFIG_FILE    = os.path.join(_PROJ, "garmin_config.json")
LOCATIONS_FILE = os.path.join(_BASE, "sup_locations.json")

MONTHS_HEB = {
    1:"ינואר", 2:"פברואר", 3:"מרץ",    4:"אפריל",
    5:"מאי",   6:"יוני",   7:"יולי",   8:"אוגוסט",
    9:"ספטמבר",10:"אוקטובר",11:"נובמבר",12:"דצמבר"
}

# ── אימות גרמין ──────────────────────────────────────────────────────────────
def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}

def connect_garmin():
    """
    התחברות חכמה לגרמין:
    1. אם קיים טוקן שמור — נשתמש בו (ללא login חדש)
    2. אחרת — login עם email+password ושמירת טוקן
    """
    token_dir = os.path.join(_PROJ, ".garth_tokens_1")

    # נסה טוקן שמור קודם
    if os.path.exists(token_dir):
        try:
            api = garminconnect.Garmin()
            api.login(tokenstore=token_dir)
            print("  [Garmin] כניסה עם טוקן שמור")
            return api
        except Exception:
            pass  # טוקן פג — נחדש

    # כניסה עם פרטים
    creds = load_config()
    email    = creds.get("email")    or input("אימייל גרמין קונקט: ").strip()
    password = creds.get("password") or getpass.getpass("סיסמה: ")
    try:
        api = garminconnect.Garmin(email, password)
        api.login()
        # שמור טוקן לשימוש הבא
        os.makedirs(token_dir, exist_ok=True)
        api.garth.dump(token_dir)
        print("  [Garmin] כניסה בוצעה, טוקן נשמר")
        # שמור פרטים אם אין
        if not creds:
            ans = input("שמור פרטי כניסה? (כ/ל): ").strip().lower()
            if ans == "כ":
                with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                    json.dump({"email": email, "password": password}, f)
        return api
    except garminconnect.GarminConnectAuthenticationError:
        print("שגיאת אימות — בדוק אימייל/סיסמה.")
        sys.exit(1)
    except Exception as e:
        if "429" in str(e):
            print("גרמין חסם זמנית (Too Many Requests) — נסה שוב עוד 20 דקות.")
        else:
            print(f"שגיאת חיבור: {e}")
        sys.exit(1)

# ── זיהוי מיקום מ-GPS ────────────────────────────────────────────────────────
def haversine_m(lat1, lon1, lat2, lon2):
    """מרחק בין שתי נקודות GPS במטרים"""
    R = 6_371_000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlam/2)**2
    return 2 * R * math.asin(math.sqrt(a))

def load_locations():
    if os.path.exists(LOCATIONS_FILE):
        with open(LOCATIONS_FILE, encoding="utf-8") as f:
            return json.load(f)
    return []

def save_locations(locs):
    with open(LOCATIONS_FILE, "w", encoding="utf-8") as f:
        json.dump(locs, f, ensure_ascii=False, indent=2)

def detect_location(start_lat, start_lon):
    """זהה מיקום לפי קואורדינטות התחלה"""
    locations = load_locations()
    if not locations:
        return None
    best_name, best_dist = None, float("inf")
    for loc in locations:
        d = haversine_m(start_lat, start_lon, loc["lat"], loc["lon"])
        if d < best_dist:
            best_dist = d
            best_name = loc["name"]
    # אם בתוך רדיוס 1 ק"מ — ודאי
    if best_dist <= 1000:
        return best_name
    return None

def setup_locations(api):
    """אשף הגדרת מיקומים — מריץ פעם אחת"""
    print("\n── הגדרת מיקומים ──────────────────────────────")
    print("נשלוף את 10 האימונים האחרונים מגרמין לזיהוי נקודות ההתחלה שלך.\n")

    activities = api.get_activities(0, 20)
    sup = [a for a in activities if "stand_up_paddle" in a.get("activityType", {}).get("typeKey","").lower()][:10]

    if not sup:
        print("לא נמצאו אימוני SUP — הזן מיקומים ידנית.")
        return

    locations = []
    seen = set()
    for a in sup:
        lat = a.get("startLatitude")
        lon = a.get("startLongitude")
        if not lat or not lon:
            continue
        key = (round(lat,3), round(lon,3))
        if key in seen:
            continue
        seen.add(key)
        date = a["startTimeLocal"][:10]
        print(f"  אימון {date}  |  lat={lat:.4f}, lon={lon:.4f}")
        name = input("  מיקום (ים/נחל/דלג): ").strip()
        if name and name != "דלג":
            locations.append({"name": name, "lat": lat, "lon": lon})

    save_locations(locations)
    print(f"\nנשמרו {len(locations)} מיקומים ב-{LOCATIONS_FILE}")

# ── ניתוח גרף דופק — זיהוי ספרינטים ─────────────────────────────────────────
def get_hr_timeseries(api, activity_id):
    """שלוף נתוני HR לאורך זמן מהאימון"""
    try:
        details = api.get_activity_details(activity_id, maxChartSize=2000)
        descriptors = details.get("metricDescriptors", [])
        hr_idx = None
        for d in descriptors:
            if d.get("key") == "directHeartRate":
                hr_idx = d.get("metricsIndex")
                break
        if hr_idx is None:
            return []
        hr = []
        for point in details.get("activityDetailMetrics", []):
            vals = point.get("metrics", [])
            if hr_idx < len(vals) and vals[hr_idx] is not None:
                hr.append(float(vals[hr_idx]))
        return hr
    except Exception:
        return []

def smooth(data, window=10):
    """ממוצע נע"""
    arr = np.array(data, dtype=float)
    kernel = np.ones(window) / window
    return np.convolve(arr, kernel, mode="same").tolist()

def count_sprint_cycles(hr_values, min_prominence=15, min_cycles=3):
    """
    סופר מחזורי פסגה-שפל משמעותיים בגרף הדופק.
    ספרינט = עלייה מהירה + ירידה מהירה, שחוזרת על עצמה.
    """
    if len(hr_values) < 40:
        return 0

    s = smooth(hr_values, window=12)
    n = len(s)
    peaks   = [i for i in range(1, n-1) if s[i] > s[i-1] and s[i] >= s[i+1]]
    valleys = [i for i in range(1, n-1) if s[i] < s[i-1] and s[i] <= s[i+1]]

    cycles = 0
    for p in peaks:
        prev_v = [v for v in valleys if v < p]
        next_v = [v for v in valleys if v > p]
        if prev_v and next_v:
            rise = s[p] - s[max(prev_v)]   # עלייה לפסגה
            drop = s[p] - s[min(next_v)]   # ירידה מהפסגה
            if rise >= min_prominence and drop >= min_prominence:
                cycles += 1

    return cycles

# ── זיהוי סוג אימון ───────────────────────────────────────────────────────────
def hms_to_sec(t):
    """המר מחרוזת H:MM:SS לשניות"""
    if not t or t in ("0:00", ""):
        return 0
    parts = str(t).split(":")
    parts = [int(p) for p in parts]
    if len(parts) == 3:
        return parts[0]*3600 + parts[1]*60 + parts[2]
    return parts[0]*60 + parts[1]

def detect_training_type(distance_km, duration_sec, z4_str, z5_str, hr_values):
    """
    זיהוי סוג אימון אוטומטי + רמת ביטחון (0-1).
    סדר עדיפויות:
      1. אירובי ארוך — מרחק > 11 ק"מ AND זמן > 90 דקות
      2. ספרינטים    — גרף דופק מציג ≥3 מחזורי פסגה-שפל (≥15 bpm)
      3. טמפו        — Z4 > 20 דקות
      4. אירובי      — ברירת מחדל
    """
    duration_min = duration_sec / 60
    z4_sec = hms_to_sec(z4_str)
    z5_sec = hms_to_sec(z5_str)

    # 1. אירובי ארוך — מרחק > 11 ק"מ (ללא תנאי זמן)
    if distance_km > 11:
        return "אירובי ארוך", 0.95

    # 2. ספרינטים — Z5 ברור, או גרף דופק עם מחזורים
    if z5_sec > 30:
        return "ספרינטים", 0.90

    cycles = count_sprint_cycles(hr_values) if hr_values else 0
    if cycles >= 3:
        conf = min(0.70 + cycles * 0.05, 0.95)
        return "ספרינטים", conf

    # 3. טמפו
    if z4_sec > 1200:   # 20 דקות
        return "טמפו", 0.88

    # 4. אירובי
    return "אירובי", 0.80

# ── השוואה ל-5 אימונים אחרונים ───────────────────────────────────────────────
def get_prev_month_stats(training_type, location, current_date_str, n=5):
    """
    שלוף ממוצע של N האימונים האחרונים באותו סוג + מיקום (לא כולל האימון הנוכחי).
    """
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["נתוני אימונים"]

        cur = datetime.strptime(current_date_str, "%d.%m.%Y")

        rows = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            date_val, type_val, loc_val, _, _, speed, hr, _, dps = (row + (None,)*9)[:9]
            if not date_val or not type_val:
                continue
            try:
                d = datetime.strptime(str(date_val), "%d.%m.%Y")
            except Exception:
                continue
            # אותו סוג + מיקום + לפני האימון הנוכחי
            if type_val == training_type and loc_val == location and d < cur:
                rows.append({"date": d, "speed": speed, "hr": hr, "dps": dps})

        wb.close()

        # מיין לפי תאריך יורד — קח N אחרונים
        rows.sort(key=lambda r: r["date"], reverse=True)
        last_n = rows[:n]

        if not last_n:
            return None

        speeds = [float(r["speed"]) for r in last_n if r["speed"]]
        hrs    = [float(r["hr"])    for r in last_n if r["hr"]]
        dpss   = [float(r["dps"])   for r in last_n if r["dps"]]

        if not speeds:
            return None

        oldest = last_n[-1]["date"].strftime("%d.%m.%y")
        newest = last_n[0]["date"].strftime("%d.%m.%y")
        label  = f"{oldest} – {newest}" if len(last_n) > 1 else newest

        return {
            "month": label,
            "count": len(last_n),
            "speed": round(sum(speeds)/len(speeds), 1),
            "hr":    round(sum(hrs)/len(hrs))     if hrs  else None,
            "dps":   round(sum(dpss)/len(dpss), 2) if dpss else None,
        }
    except Exception:
        return None

def trend_arrow(current, prev):
    """↑ / → / ↓ לפי הפרש"""
    if prev is None:
        return ""
    diff = current - prev
    if diff > 0.05:  return "↑"
    if diff < -0.05: return "↓"
    return "→"

def send_toast(data, prev_stats):
    """שלח Windows Toast Notification עם סיכום + השוואה"""
    if not TOAST_AVAILABLE:
        print("  [Toast לא זמין — התקן win10toast]")
        return

    title = f"🏄 SUP | {data['date']} — {data['type']} {data['location']}"

    # שורה 1: נתוני האימון
    lines = [
        f"מרחק: {data['distance']} ק\"מ  |  זמן: {data['duration']}  |  מהירות: {data['speed']} קמ\"ש",
        f"דופק: {data['avg_hr']} bpm  |  DPS: {data['dps']} מ'  |  SPM: {data['spm']}",
    ]

    # שורה 2: השוואה לחודש קודם
    if prev_stats:
        sp_arr = trend_arrow(data['speed'],   prev_stats['speed'])
        hr_arr = trend_arrow(prev_stats['hr'] or data['avg_hr'], data['avg_hr']) if prev_stats['hr'] else ""
        dps_arr= trend_arrow(data['dps'],     prev_stats['dps'] or 0)
        lines.append(
            f"vs {prev_stats['month']} ({prev_stats['count']} אימונים):"
        )
        lines.append(
            f"מהירות {sp_arr} {prev_stats['speed']}  |  "
            f"דופק {hr_arr} {prev_stats['hr'] or '-'}  |  "
            f"DPS {dps_arr} {prev_stats['dps'] or '-'}"
        )
    else:
        lines.append("אין נתוני השוואה לחודש הקודם")

    body = "\n".join(lines)

    try:
        toast = Notification(
            app_id="SUP Tracker",
            title=title,
            msg=body,
            duration="long",
        )
        toast.set_audio(audio.Default, loop=False)
        toast.show()
        print(f"  🔔 נשלחה התראה ל-Windows")
    except Exception as e:
        print(f"  שגיאה בשליחת התראה: {e}")

# ── שליחת אימייל עם דוח HTML ─────────────────────────────────────────────────
def send_email_report(html_path, data, prev_stats):
    """שלח את דוח ה-HTML כגוף האימייל ל-Gmail"""
    GMAIL_USER = "maxim.maxster@gmail.com"
    GMAIL_APP_PASSWORD = "wueohtjazghiiqwq"  # App Password (ללא רווחים)
    TO_EMAIL = "maxim.maxster@gmail.com"

    try:
        with open(html_path, encoding="utf-8") as f:
            html_content = f.read()

        subject = f"🏄 SUP | {data['date']} — {data['type']} {data['location']} | {data['distance']} ק\"מ"

        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = GMAIL_USER
        msg["To"] = TO_EMAIL
        msg.attach(MIMEText(html_content, "html", "utf-8"))

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            server.sendmail(GMAIL_USER, TO_EMAIL, msg.as_string())

        print(f"  [Email] נשלח ל-{TO_EMAIL}")
    except Exception as e:
        print(f"  [Email] שגיאה: {e}")

# ── עיבוד נתוני גרמין ────────────────────────────────────────────────────────
def seconds_to_hms(sec):
    if not sec:
        return "0:00"
    h = int(sec // 3600)
    m = int((sec % 3600) // 60)
    s = int(sec % 60)
    return f"{h}:{m:02d}:{s:02d}" if h else f"{m}:{s:02d}"

def get_zone_time(zones, zone_number):
    if not zones:
        return "0:00"
    for z in zones:
        if z.get("zoneNumber") == zone_number:
            return seconds_to_hms(z.get("secsInZone", 0))
    return "0:00"

def format_activity(activity, zones):
    start_dt = datetime.fromisoformat(activity["startTimeLocal"].replace("Z",""))
    dist_km  = round(activity.get("distance", 0) / 1000, 2)
    dur_sec  = activity.get("duration", 0)
    speed    = round(dist_km / (dur_sec / 3600), 1) if dur_sec else 0
    avg_hr   = int(activity.get("averageHR", 0) or 0)
    # SPM — חשב מסך משיכות / דקות
    total_strokes = activity.get("strokes", 0) or 0
    avg_spm = int(round(total_strokes / (dur_sec / 60))) if dur_sec and total_strokes else 0

    # DPS — avgStrokeDistance מגיע בסנטימטרים → המר למטרים
    avg_stroke_dist = activity.get("avgStrokeDistance", 0) or 0
    avg_dps = round(avg_stroke_dist / 100, 2) if avg_stroke_dist > 10 else round(avg_stroke_dist, 2)

    return {
        "date":       start_dt.strftime("%d.%m.%Y"),
        "start_time": start_dt.strftime("%H:%M"),
        "distance":   dist_km,
        "duration":   seconds_to_hms(dur_sec),
        "dur_sec":    dur_sec,
        "speed":      speed,
        "avg_hr":     avg_hr,
        "spm":        avg_spm,
        "dps":        avg_dps,
        "z1":         get_zone_time(zones, 1),
        "z2":         get_zone_time(zones, 2),
        "z3":         get_zone_time(zones, 3),
        "z4":         get_zone_time(zones, 4),
        "z5":         get_zone_time(zones, 5),
        "start_dt":   start_dt,
    }

# ── בדיקת כפילות ─────────────────────────────────────────────────────────────
def is_already_logged(date_str):
    """בדוק אם התאריך כבר קיים באקסל — למניעת כפילות"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["נתוני אימונים"]
        for row in ws.iter_rows(min_row=3, max_col=1, values_only=True):
            if row[0] == date_str:
                wb.close()
                return True
        wb.close()
    except Exception:
        pass
    return False

# ── עדכון אקסל ───────────────────────────────────────────────────────────────
def get_last_date_from_excel():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["נתוני אימונים"]
    last = None
    for row in ws.iter_rows(min_row=3, max_col=1, values_only=True):
        if row[0]:
            last = row[0]
    wb.close()
    return last

def add_training_row(ws, data):
    new_row = ws.max_row + 1
    for col in range(1, 13):
        src = ws.cell(row=new_row-1, column=col)
        tgt = ws.cell(row=new_row,   column=col)
        if src.has_style:
            tgt.font = copy(src.font); tgt.fill = copy(src.fill)
            tgt.border = copy(src.border); tgt.alignment = copy(src.alignment)
            tgt.number_format = src.number_format

    for col, val in enumerate([
        data["date"], data["type"], data["location"],
        data["distance"], data["duration"], data["speed"],
        data["avg_hr"], data["spm"], data["dps"],
        data["z3"], data["z4"], data["z5"],
    ], start=1):
        ws.cell(row=new_row, column=col).value = val

def ensure_month_in_summary(ws, month_heb, pattern):
    for row in ws.iter_rows(min_row=3, max_col=1, values_only=True):
        if row[0] == month_heb:
            return
    insert_at = ws.max_row + 1
    for i, row in enumerate(ws.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
        if row[0] and "🏆" in str(row[0]):
            insert_at = i
            break
    ws.insert_rows(insert_at, 7)
    types = ["אירובי", "אירובי ארוך", "טמפו", "ספרינטים", "סה\"כ חודש"]
    for i, t in enumerate(types):
        row = insert_at + 2 + i
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row <= row <= mr.max_row:
                ws.unmerge_cells(str(mr))
        ws.cell(row=row, column=1).value = month_heb
        ws.cell(row=row, column=2).value = t
        if t == "סה\"כ חודש":
            ws.cell(row=row, column=3).value = f"=IFERROR(COUNTIFS('נתוני אימונים'!A:A,\"{pattern}\"),0)"
            ws.cell(row=row, column=4).value = f"=IFERROR(AVERAGEIFS('נתוני אימונים'!F:F,'נתוני אימונים'!A:A,\"{pattern}\"),\"\")"
            ws.cell(row=row, column=5).value = f"=IFERROR(AVERAGEIFS('נתוני אימונים'!G:G,'נתוני אימונים'!A:A,\"{pattern}\"),\"\")"
            ws.cell(row=row, column=6).value = f"=IFERROR(AVERAGEIFS('נתוני אימונים'!I:I,'נתוני אימונים'!A:A,\"{pattern}\"),\"\")"
            ws.cell(row=row, column=7).value = f"=IFERROR(SUMIFS('נתוני אימונים'!D:D,'נתוני אימונים'!A:A,\"{pattern}\"),\"\")"
        else:
            ws.cell(row=row, column=3).value = f"=IFERROR(COUNTIFS('נתוני אימונים'!A:A,\"{pattern}\",'נתוני אימונים'!B:B,\"{t}\"),0)"
            ws.cell(row=row, column=4).value = f"=IFERROR(AVERAGEIFS('נתוני אימונים'!F:F,'נתוני אימונים'!A:A,\"{pattern}\",'נתוני אימונים'!B:B,\"{t}\"),\"\")"
            ws.cell(row=row, column=5).value = f"=IFERROR(AVERAGEIFS('נתוני אימונים'!G:G,'נתוני אימונים'!A:A,\"{pattern}\",'נתוני אימונים'!B:B,\"{t}\"),\"\")"
            ws.cell(row=row, column=6).value = f"=IFERROR(AVERAGEIFS('נתוני אימונים'!I:I,'נתוני אימונים'!A:A,\"{pattern}\",'נתוני אימונים'!B:B,\"{t}\"),\"\")"
            ws.cell(row=row, column=7).value = f"=IFERROR(SUMIFS('נתוני אימונים'!D:D,'נתוני אימונים'!A:A,\"{pattern}\",'נתוני אימונים'!B:B,\"{t}\"),\"\")"

def ensure_month_in_graphs(ws, month_heb, pattern):
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] == month_heb:
            return
    r = ws.max_row + 1
    ws.cell(row=r, column=1).value = month_heb
    ws.cell(row=r, column=2).value = f"=IFERROR(AVERAGEIFS('נתוני אימונים'!F:F,'נתוני אימונים'!A:A,\"{pattern}\"),0)"
    ws.cell(row=r, column=3).value = f"=IFERROR(AVERAGEIFS('נתוני אימונים'!G:G,'נתוני אימונים'!A:A,\"{pattern}\"),0)"
    ws.cell(row=r, column=4).value = f"=IFERROR(AVERAGEIFS('נתוני אימונים'!I:I,'נתוני אימונים'!A:A,\"{pattern}\"),0)"
    ws.cell(row=r, column=5).value = f"=IFERROR(B{r}/C{r},0)"

# ── היסטוריית אימונים ─────────────────────────────────────────────────────────
def get_session_history(training_type, location, limit=5):
    """שלוף N אימונים אחרונים מאותו סוג + מיקום מהאקסל"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["נתוני אימונים"]
        rows = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            date_val, type_val, loc_val, dist, dur, speed, hr, spm, dps = (row + (None,)*9)[:9]
            if not date_val:
                continue
            if type_val == training_type and loc_val == location:
                rows.append({
                    "date": str(date_val),
                    "distance": dist,
                    "duration": dur,
                    "speed": speed,
                    "avg_hr": hr,
                    "dps": dps,
                })
        wb.close()
        # מיון לפי תאריך — האחרונים ראשונים
        def parse_date(r):
            try:
                return datetime.strptime(r["date"], "%d.%m.%Y")
            except Exception:
                return datetime.min
        rows.sort(key=parse_date, reverse=True)
        return rows[:limit]
    except Exception:
        return []

# ── יצירת דוח HTML ────────────────────────────────────────────────────────────
def generate_html_report(data, prev_stats, history, open_browser=False):
    """צור דוח HTML מעוצב ופתח בדפדפן"""

    # יום בשבוע בעברית
    try:
        start_dt = datetime.strptime(data["date"], "%d.%m.%Y")
        day_name = DAYS_HEB[start_dt.weekday()]
        date_display = f"יום {day_name}, {data['date']} | {data.get('start_time','')}"
    except Exception:
        date_display = data["date"]

    # זונות דופק — חישוב אחוזים
    def hms_secs(t):
        if not t or t in ("0:00",""):
            return 0
        parts = [int(x) for x in str(t).split(":")]
        if len(parts) == 3:
            return parts[0]*3600 + parts[1]*60 + parts[2]
        return parts[0]*60 + parts[1]

    total_sec = data["dur_sec"]
    z1_s = hms_secs(data.get("z1", ""))
    z2_s = hms_secs(data.get("z2", ""))
    z3_s = hms_secs(data["z3"])
    z4_s = hms_secs(data["z4"])
    z5_s = hms_secs(data["z5"])
    # אם Z1/Z2 לא ידועים — חשב לפי שאריות
    if not z1_s and not z2_s:
        z2_s = max(0, total_sec - z3_s - z4_s - z5_s - 60)
        z1_s = max(0, total_sec - z2_s - z3_s - z4_s - z5_s)

    # חשב רוחב בפיקסלים — יחסית לzone הגדולה (מתוך 380px)
    max_zone = max(z1_s, z2_s, z3_s, z4_s, z5_s, 1)
    BAR_MAX_PX = 380

    def px(s):
        return max(6, round(s / max_zone * BAR_MAX_PX)) if s else 0

    def pct_abs(s):
        return round(s / total_sec * 100) if total_sec else 0

    def fmt_pct(s):
        p = pct_abs(s)
        return f"{p}%" if p >= 4 else ""

    # צבעי סוג אימון
    type_colors = {
        "אירובי":       ("#1b5e20", "#2e7d32", "#a5d6a7"),
        "אירובי ארוך":  ("#004d40", "#00695c", "#80cbc4"),
        "טמפו":         ("#e65100", "#ef6c00", "#ffcc80"),
        "ספרינטים":     ("#b71c1c", "#c62828", "#ef9a9a"),
    }
    bg1, bg2, text_c = type_colors.get(data["type"], ("#1b5e20","#2e7d32","#a5d6a7"))

    # השוואה
    def delta_html(curr, prev_val, reverse=False):
        """reverse=True אומר שירידה היא טובה (כמו דופק)"""
        if prev_val is None:
            return ""
        try:
            diff = float(curr) - float(prev_val)
            pct_diff = abs(diff) / float(prev_val) * 100 if prev_val else 0
        except Exception:
            return ""
        if abs(diff) < 0.05:
            color, arrow = "#ffa726", "→"
            label = "ללא שינוי"
        elif (diff > 0) != reverse:
            color, arrow = "#66bb6a", "↑"
            label = f"+{pct_diff:.1f}%"
        else:
            color, arrow = "#ef5350", "↓"
            label = f"-{pct_diff:.1f}%"
        return f'<div class="delta" style="color:{color}">{arrow} {label}</div>'

    compare_html = ""
    if prev_stats:
        compare_html = f"""
        <div class="section">
          <div class="section-title">📊 ממוצע {prev_stats['count']} אימונים אחרונים — {data['type']} {data['location']} ({prev_stats['month']})</div>
          <div class="compare-grid">
            <div class="compare-card">
              <div class="clabel">מהירות (קמ"ש)</div>
              <div class="curr">{data['speed']}</div>
              <div class="prev">{prev_stats['month'][:3]}: {prev_stats['speed']}</div>
              {delta_html(data['speed'], prev_stats['speed'])}
            </div>
            <div class="compare-card">
              <div class="clabel">דופק ממוצע</div>
              <div class="curr">{data['avg_hr']}</div>
              <div class="prev">{prev_stats['month'][:3]}: {prev_stats.get('hr') or '-'}</div>
              {delta_html(data['avg_hr'], prev_stats.get('hr'), reverse=True)}
            </div>
            <div class="compare-card">
              <div class="clabel">DPS (מטר)</div>
              <div class="curr">{data['dps']}</div>
              <div class="prev">{prev_stats['month'][:3]}: {prev_stats.get('dps') or '-'}</div>
              {delta_html(data['dps'], prev_stats.get('dps'))}
            </div>
          </div>
        </div>"""
    else:
        compare_html = """
        <div class="section" style="text-align:center;color:#546e7a;padding:14px">
          אין נתוני השוואה לחודש הקודם עבור סוג ומיקום זה
        </div>"""

    # טבלת היסטוריה
    hist_rows = ""
    for i, h in enumerate(history):
        cls = 'class="highlight-row"' if i == 0 else ""
        today_tag = '<span class="today-tag">← היום</span>' if i == 0 else ""
        hist_rows += f"""
            <tr {cls}>
              <td>{h['date']}{today_tag}</td>
              <td>{h.get('distance','')}</td>
              <td>{h.get('duration','')}</td>
              <td>{h.get('speed','')}</td>
              <td>{h.get('avg_hr','')}</td>
              <td>{h.get('dps','')}</td>
            </tr>"""

    hist_section = ""
    if hist_rows:
        hist_section = f"""
        <div class="section">
          <div class="section-title">📋 אימונים אחרונים — {data['type']} {data['location']}</div>
          <table>
            <thead><tr>
              <th>תאריך</th><th>מרחק</th><th>זמן</th><th>מהירות</th><th>דופק</th><th>DPS</th>
            </tr></thead>
            <tbody>{hist_rows}</tbody>
          </table>
        </div>"""

    html = f"""<!DOCTYPE html>
<html dir="rtl" lang="he">
<head>
  <meta charset="UTF-8">
  <title>SUP | {data['date']} — {data['type']} {data['location']}</title>
  <style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #0f1923; color: #e0e0e0; padding: 20px; direction: rtl; }}
    .container {{ max-width: 680px; margin: 0 auto; }}
    .header {{ background: linear-gradient(135deg, #1a3a5c, #0d2137); border-radius: 16px; padding: 24px 28px; margin-bottom: 16px; border: 1px solid #1e4d7a; display: flex; align-items: center; gap: 16px; }}
    .header-icon {{ font-size: 48px; }}
    .header-text h1 {{ font-size: 22px; color: #4fc3f7; font-weight: 700; }}
    .header-text .subtitle {{ font-size: 13px; color: #78909c; margin-top: 4px; }}
    .header-text .date {{ font-size: 14px; color: #90caf9; margin-top: 6px; }}
    .cards-row {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; margin-bottom: 16px; }}
    .card {{ background: #1a2a3a; border-radius: 12px; padding: 16px; text-align: center; border: 1px solid #1e3a55; }}
    .card .label {{ font-size: 11px; color: #78909c; margin-bottom: 6px; text-transform: uppercase; }}
    .card .value {{ font-size: 26px; font-weight: 700; color: #4fc3f7; }}
    .card .unit {{ font-size: 12px; color: #546e7a; margin-top: 2px; }}
    .section {{ background: #1a2a3a; border-radius: 12px; padding: 20px; margin-bottom: 16px; border: 1px solid #1e3a55; }}
    .section-title {{ font-size: 14px; color: #78909c; margin-bottom: 14px; font-weight: 600; text-transform: uppercase; }}
    .zone-row {{ display: flex; align-items: center; margin-bottom: 10px; gap: 10px; }}
    .zone-label {{ width: 30px; font-size: 13px; color: #90a4ae; text-align: right; flex-shrink: 0; font-weight: 600; }}
    .zone-bar-bg {{ width: 380px; background: #0d1e2e; border-radius: 6px; height: 20px; overflow: hidden; flex-shrink: 0; }}
    .zone-bar {{ height: 20px; border-radius: 6px; display: inline-flex; align-items: center; justify-content: flex-end; padding: 0 8px; font-size: 11px; color: white; font-weight: 700; min-width: 6px; box-sizing: border-box; }}
    .zone-time {{ width: 55px; font-size: 12px; color: #90caf9; text-align: left; flex-shrink: 0; }}
    .type-banner {{ background: linear-gradient(135deg, {bg1}, {bg2}); border-radius: 10px; padding: 12px 18px; margin-bottom: 16px; display: flex; align-items: center; justify-content: space-between; border: 1px solid {bg2}; }}
    .type-banner .type-name {{ font-size: 18px; font-weight: 700; color: {text_c}; }}
    .type-banner .type-loc {{ font-size: 13px; color: {text_c}; opacity: 0.8; margin-top:3px }}
    .type-banner .type-icon {{ font-size: 32px; }}
    .compare-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 8px; }}
    .compare-card {{ background: #0d1e2e; border-radius: 10px; padding: 14px 12px; text-align: center; }}
    .compare-card .clabel {{ font-size: 11px; color: #546e7a; margin-bottom: 8px; }}
    .compare-card .curr {{ font-size: 22px; font-weight: 700; color: #e0e0e0; }}
    .compare-card .prev {{ font-size: 11px; color: #546e7a; margin-top: 4px; }}
    .compare-card .delta {{ font-size: 13px; font-weight: 600; margin-top: 6px; }}
    .chips-row {{ display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 16px; }}
    .pace-chip {{ display: inline-flex; align-items: center; gap: 6px; background: #0d2137; border: 1px solid #1e4d7a; border-radius: 8px; padding: 8px 14px; font-size: 13px; color: #90caf9; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    thead th {{ background: #0d1e2e; color: #546e7a; padding: 10px 8px; font-weight: 600; font-size: 11px; text-align: center; }}
    tbody tr {{ border-bottom: 1px solid #1a2a3a; }}
    tbody tr:hover {{ background: #1e3040; }}
    tbody td {{ padding: 10px 8px; color: #90a4ae; text-align: center; }}
    .highlight-row td {{ color: #4fc3f7 !important; font-weight: 600; }}
    .today-tag {{ font-size: 10px; color: #4fc3f7; margin-right: 4px; }}
    .footer {{ text-align: center; margin-top: 16px; font-size: 11px; color: #37474f; }}
  </style>
</head>
<body>
<div class="container">
  <div class="header">
    <div class="header-icon">🏄</div>
    <div class="header-text">
      <h1>סיכום אימון SUP</h1>
      <div class="date">{date_display}</div>
      <div class="subtitle">עודכן אוטומטית מ-Garmin Connect</div>
    </div>
  </div>

  <div class="type-banner">
    <div>
      <div class="type-name">{data['type']}</div>
      <div class="type-loc">📍 {data['location']}</div>
    </div>
    <div class="type-icon">🌊</div>
  </div>

  <div class="cards-row">
    <div class="card"><div class="label">מרחק</div><div class="value">{data['distance']}</div><div class="unit">ק"מ</div></div>
    <div class="card"><div class="label">זמן</div><div class="value">{data['duration']}</div><div class="unit">דקות</div></div>
    <div class="card"><div class="label">מהירות</div><div class="value">{data['speed']}</div><div class="unit">קמ"ש</div></div>
  </div>

  <div class="chips-row">
    <div class="pace-chip">💓 דופק ממוצע: <strong>{data['avg_hr']} bpm</strong></div>
    <div class="pace-chip">🚣 קצב: <strong>{data['spm']} SPM</strong></div>
    <div class="pace-chip">📏 DPS: <strong>{data['dps']} מ'</strong></div>
  </div>

  <div class="section">
    <div class="section-title">⏱ זמן בזונות דופק</div>
    <div class="zone-row">
      <div class="zone-label">Z1</div>
      <div class="zone-bar-bg"><div class="zone-bar" style="width:{px(z1_s)}px;background:#37474f">{fmt_pct(z1_s)}</div></div>
      <div class="zone-time">{data.get('z1') or f"{z1_s//60}:{z1_s%60:02d}"}</div>
    </div>
    <div class="zone-row">
      <div class="zone-label">Z2</div>
      <div class="zone-bar-bg"><div class="zone-bar" style="width:{px(z2_s)}px;background:#1565c0">{fmt_pct(z2_s)}</div></div>
      <div class="zone-time">{data.get('z2') or f"{z2_s//60}:{z2_s%60:02d}"}</div>
    </div>
    <div class="zone-row">
      <div class="zone-label">Z3</div>
      <div class="zone-bar-bg"><div class="zone-bar" style="width:{px(z3_s)}px;background:#2e7d32">{fmt_pct(z3_s)}</div></div>
      <div class="zone-time">{data['z3']}</div>
    </div>
    <div class="zone-row">
      <div class="zone-label">Z4</div>
      <div class="zone-bar-bg"><div class="zone-bar" style="width:{px(z4_s)}px;background:#e65100">{fmt_pct(z4_s)}</div></div>
      <div class="zone-time">{data['z4']}</div>
    </div>
    <div class="zone-row">
      <div class="zone-label">Z5</div>
      <div class="zone-bar-bg"><div class="zone-bar" style="width:{px(z5_s)}px;background:#b71c1c">{fmt_pct(z5_s)}</div></div>
      <div class="zone-time">{data['z5']}</div>
    </div>
  </div>

  {compare_html}
  {hist_section}

  <div class="footer">נוצר אוטומטית על ידי SUP Tracker • Garmin Connect • {data['date']}</div>
</div>
</body>
</html>"""

    # שמור לקובץ בתיקיית הפרויקט ופתח בדפדפן
    script_dir  = os.path.dirname(os.path.abspath(__file__))
    fname       = f"sup_report_{data['date'].replace('.','_')}.html"
    report_path = os.path.join(script_dir, fname)
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(html)

    # שמור גם בשם קבוע latest.html
    latest_path = os.path.join(script_dir, "latest.html")
    with open(latest_path, "w", encoding="utf-8") as f:
        f.write(html)

    if open_browser:
        url = f"http://localhost:8765/{fname}"
        try:
            urllib.request.urlopen(url, timeout=1)
            webbrowser.open(url)
        except Exception:
            webbrowser.open(f"file:///{report_path.replace(os.sep, '/')}")
    print(f"  [HTML] נוצר: {report_path}")
    return report_path

# ── לוג תוצאות ───────────────────────────────────────────────────────────────
def log_result(date, location, training_type, confidence, location_known):
    try:
        log_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "activity_log.txt")
        with open(log_file, "a", encoding="utf-8") as f:
            loc_flag = "OK" if location_known else "?"
            f.write(f"{datetime.now():%Y-%m-%d %H:%M} | {date} | {location} {loc_flag} | "
                    f"{training_type} ({confidence:.0%})\n")
    except Exception:
        pass  # לוג לא חיוני — לא קורסים בגללו

# ── ראשי ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--setup", action="store_true", help="הגדר מיקומים ידועים")
    args = parser.parse_args()

    print("=" * 55)
    print("  מעקב אימוני SUP — אוטומטי מלא")
    print("=" * 55)

    # בדוק חיבור אינטרנט — ממתין עד 30 דקות אם צריך
    if not wait_for_internet(max_wait_min=30):
        sys.exit(0)

    api = connect_garmin()

    if args.setup:
        setup_locations(api)
        return

    # בדוק שקובץ מיקומים קיים
    if not os.path.exists(LOCATIONS_FILE):
        print("\nטרם הוגדרו מיקומים. מריץ אשף הגדרה...")
        setup_locations(api)

    # תאריך אחרון באקסל
    last_str = get_last_date_from_excel()
    last_date = datetime.strptime(last_str, "%d.%m.%Y") if last_str else datetime.now() - timedelta(days=90)
    print(f"תאריך אחרון באקסל: {last_str or 'לא נמצא'}")

    start = (last_date + timedelta(days=1)).strftime("%Y-%m-%d")
    end   = datetime.now().strftime("%Y-%m-%d")
    print(f"מחפש אימוני SUP: {start} → {end}...")

    try:
        all_activities = api.get_activities_by_date(start, end, "stand_up_paddleboarding")
    except Exception:
        # fallback — שלוף הכל וסנן ידנית
        all_activities = api.get_activities_by_date(start, end)

    activities = [a for a in all_activities
                  if "paddle" in a.get("activityType", {}).get("typeKey", "").lower()]

    if not activities:
        print("אין אימוני SUP חדשים.")
        return

    activities.sort(key=lambda a: a["startTimeLocal"])
    print(f"נמצאו {len(activities)} אימון(ים) חדש(ים).\n")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_data    = wb["נתוני אימונים"]
    ws_monthly = wb["סיכום חודשי"]
    ws_graphs  = wb["גרפים"]

    added = 0
    for act in activities:
        act_id   = act["activityId"]
        start_dt = datetime.fromisoformat(act["startTimeLocal"].replace("Z",""))

        print(f"─── {start_dt.strftime('%d.%m.%Y')} ───────────────────────────────")

        # HR zones
        try:
            zones = api.get_activity_hr_in_timezones(act_id)
        except Exception:
            zones = []

        # HR time series לזיהוי ספרינטים
        hr_ts = get_hr_timeseries(api, act_id)

        # נסה לקבל פרטים נוספים (SPM, DPS) מהאימון
        try:
            act_detail = api.get_activity(act_id)
            for key in ["averageStrokeCadence", "avgStrokeCadence", "avgStrideLength", "avgStrokeDistance"]:
                if act_detail.get(key) and not act.get(key):
                    act[key] = act_detail[key]
        except Exception:
            pass

        data = format_activity(act, zones)

        # ── זיהוי מיקום ──
        lat = act.get("startLatitude")
        lon = act.get("startLongitude")
        location_known = False
        if lat and lon:
            detected_loc = detect_location(lat, lon)
            if detected_loc:
                data["location"] = detected_loc
                location_known = True
            else:
                # מיקום לא מוכר — שאל
                print(f"  מיקום לא מוכר (lat={lat:.4f}, lon={lon:.4f})")
                ans = input("  מיקום (ים/נחל): ").strip()
                data["location"] = ans
                # שמור מיקום חדש
                locs = load_locations()
                locs.append({"name": ans, "lat": lat, "lon": lon})
                save_locations(locs)
        else:
            data["location"] = input("  מיקום (ים/נחל): ").strip()

        # ── זיהוי סוג אימון ──
        training_type, confidence = detect_training_type(
            data["distance"], data["dur_sec"],
            data["z4"], data["z5"], hr_ts
        )
        data["type"] = training_type

        # הדפס סיכום
        sprint_cycles = count_sprint_cycles(hr_ts) if hr_ts else "N/A"
        print(f"  מרחק: {data['distance']} ק\"מ  |  זמן: {data['duration']}  |  "
              f"מהירות: {data['speed']} קמ\"ש")
        print(f"  דופק: {data['avg_hr']} bpm  |  SPM: {data['spm']}  |  DPS: {data['dps']} מ'")
        print(f"  Z3: {data['z3']}  |  Z4: {data['z4']}  |  Z5: {data['z5']}  |  "
              f"מחזורי ספרינט: {sprint_cycles}")
        print(f"  מיקום:  {data['location']} {'✓' if location_known else '(חדש)'}")
        print(f"  סוג:    {data['type']} (ביטחון: {confidence:.0%})")

        # השוואה לחודש קודם (לפני שמירה לאקסל)
        prev_stats = get_prev_month_stats(data["type"], data["location"], data["date"])

        # הוסף לאקסל
        month_heb = f"{MONTHS_HEB[start_dt.month]} {start_dt.year}"
        pattern   = f"*.{start_dt.month:02d}.{start_dt.year}*"
        add_training_row(ws_data, data)
        ensure_month_in_summary(ws_monthly, month_heb, pattern)
        ensure_month_in_graphs(ws_graphs, month_heb, pattern)

        # בדוק כפילות — אם כבר רשום, דלג על האימייל
        already_logged = is_already_logged(data["date"])

        # יצור דוח HTML ושלח באימייל בלבד
        # בנה history: האימון הנוכחי + 4 האחרונים מהאקסל (ללא כפילות של היום)
        history_from_excel = get_session_history(data["type"], data["location"], limit=5)
        history_from_excel = [h for h in history_from_excel if h["date"] != data["date"]][:4]
        current_entry = {
            "date": data["date"], "distance": data["distance"],
            "duration": data["duration"], "speed": data["speed"],
            "avg_hr": data["avg_hr"], "dps": data["dps"],
        }
        history = [current_entry] + history_from_excel
        report_path = generate_html_report(data, prev_stats, history, open_browser=False)
        if not already_logged:
            send_email_report(report_path, data, prev_stats)
        else:
            print("  [Email] כבר נשלח קודם — מדלג")
        log_result(data["date"], data["location"], data["type"], confidence, location_known)

        added += 1
        print()

    if added:
        # נסה לשמור — אם נעול, חכה וחזור
        for attempt in range(10):
            try:
                wb.save(EXCEL_FILE)
                print(f"נשמר! {added} אימון(ים) הוזנו ל-{EXCEL_FILE}")
                break
            except PermissionError:
                if attempt < 9:
                    print(f"  האקסל נעול — ממתין 2 דקות... (ניסיון {attempt+1}/10)")
                    time.sleep(120)
                else:
                    print("  האקסל נעול — לא ניתן לשמור. סגור את הקובץ וצור קשר.")
    else:
        print("לא בוצעו שינויים.")

if __name__ == "__main__":
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    main()
